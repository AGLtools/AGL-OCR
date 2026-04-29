"""Compare reference Excel (équipe d'intégration) vs notre export MIDAS."""
import openpyxl
from pathlib import Path

REF = Path(r"testing_pdf/Look for comparison/Feuille Export.xlsx")
MINE = Path(r"testing_pdf/Look for comparison/MIDAS_MANIFESTE CMA CGM LEBU 0BAN6N1MA.xlsx")

ref_wb = openpyxl.load_workbook(REF, data_only=True, read_only=True)
mine_wb = openpyxl.load_workbook(MINE, data_only=True, read_only=True)
ref = ref_wb.active
mine = mine_wb.active

def first_n_rows(ws, n):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= n:
            break
        rows.append(list(row))
    return rows

print("=" * 80)
print(f"REF : {REF.name}  ({REF.stat().st_size/1024/1024:.1f} MB)")
print(f"  Feuille: {ref.title}")
print(f"  Sheets: {ref_wb.sheetnames}")
ref_rows = first_n_rows(ref, 4)
print(f"  Headers ({len(ref_rows[0]) if ref_rows else 0}):")
for i, v in enumerate(ref_rows[0] if ref_rows else [], 1):
    print(f"    {i:2d}. {v!r}")
for ridx, r in enumerate(ref_rows[1:], 2):
    print(f"  Row {ridx}:")
    for i, v in enumerate(r, 1):
        if v not in (None, ""):
            print(f"    {i:2d}. {v!r}")

print()
print("=" * 80)
print(f"MINE : {MINE.name}")
print(f"  Feuille: {mine.title}")
mine_rows = first_n_rows(mine, 3)
print(f"  Headers ({len(mine_rows[0]) if mine_rows else 0}):")
for i, v in enumerate(mine_rows[0] if mine_rows else [], 1):
    print(f"    {i:2d}. {v!r}")
for ridx, r in enumerate(mine_rows[1:], 2):
    print(f"  Row {ridx}:")
    for i, v in enumerate(r, 1):
        if v not in (None, ""):
            print(f"    {i:2d}. {v!r}")

print()
print("=" * 80)
print("HEADER DIFF")
def norm(x):
    return str(x).strip().lower() if x is not None else ""
ref_h = [norm(c) for c in (ref_rows[0] if ref_rows else [])]
mine_h = [norm(c) for c in (mine_rows[0] if mine_rows else [])]
ref_set = set(ref_h)
mine_set = set(mine_h)
print(f"Dans REF mais pas MINE ({len(ref_set - mine_set)}):")
for h in sorted(ref_set - mine_set):
    print(f"   - {h!r}")
print(f"Dans MINE mais pas REF ({len(mine_set - ref_set)}):")
for h in sorted(mine_set - ref_set):
    print(f"   + {h!r}")
print(f"En commun: {len(ref_set & mine_set)}")

ref_wb.close()
mine_wb.close()
