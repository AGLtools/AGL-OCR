"""Excel exporter."""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .config import EXPORTS_DIR, load_fields
from .midas_mapper import MIDAS_COLUMNS, map_rows_to_midas


class ExcelExporter:
    def __init__(self, output_path: Path | None = None):
        self.fields = load_fields()
        self.headers = ["Source File", "Template", "Page"] + [f["key"] for f in self.fields]
        self.output_path = output_path or (
            EXPORTS_DIR / f"AGL_OCR_export_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

    def export(self, rows: List[Dict]) -> Path:
        """rows: each dict has keys 'source', 'template', 'page', plus field keys.
        If rows come from the manifest parser (ManifestRow.to_dict()) they carry
        their own schema; the exporter auto-builds columns from the actual keys."""
        if not rows:
            raise ValueError("No rows to export.")

        # Detect manifest rows: they have 'bl_number' or 'vessel' keys
        # (standard OCR rows use 'BL_Number', 'Vessel_Name' etc.)
        is_manifest = any("bl_number" in r or "vessel" in r for r in rows)

        if is_manifest:
            # Build headers from the union of all keys, preserving insertion order
            all_keys: list = []
            seen: set = set()
            for r in rows:
                for k in r.keys():
                    if k not in seen:
                        all_keys.append(k)
                        seen.add(k)
            headers = all_keys
        else:
            headers = self.headers

        if self.output_path.exists():
            wb = load_workbook(self.output_path)
            ws = wb.active
            # If the file already has manifest-style columns, keep them
            existing_headers = [c.value for c in next(ws.iter_rows(max_row=1))]
            if existing_headers != headers:
                # Mismatch (e.g. old file had different schema) — start fresh sheet
                ws = wb.create_sheet(title=f"Extractions_{datetime.now():%H%M%S}")
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Extractions"
            ws.append(headers)

        for row in rows:
            ws.append([row.get(h, "") for h in headers])

        # Auto-size columns
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None),
                         default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)

        wb.save(self.output_path)
        return self.output_path

    # ============================================================
    # MIDAS export — 42 colonnes plates au format AGL
    # ============================================================
    def export_midas(self, manifest_rows: List[Dict],
                     static_overrides: Dict | None = None) -> Path:
        """Export rows directly in the MIDAS 42-column format.

        manifest_rows: list of ManifestRow.to_dict() — the rich extraction output.
        static_overrides: dict of {midas_column: forced_value} for site-specific
                          constants (e.g. {"Consignataire": "OMA CI"}).
        """
        if not manifest_rows:
            raise ValueError("Aucune ligne à exporter.")

        midas_rows = map_rows_to_midas(manifest_rows, static_overrides)

        wb = Workbook()
        ws = wb.active
        ws.title = "MIDAS"
        ws.append(MIDAS_COLUMNS)

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="1A4076")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 32

        # Data rows
        for row in midas_rows:
            ws.append([row.get(h, "") for h in MIDAS_COLUMNS])

        # Freeze header + auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Column widths
        for col_cells in ws.columns:
            values = [c.value for c in col_cells if c.value is not None]
            length = max((len(str(v)) for v in values), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 32)

        wb.save(self.output_path)
        return self.output_path
