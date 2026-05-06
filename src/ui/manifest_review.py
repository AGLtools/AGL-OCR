"""Manifest review dialog: editable table of all rows extracted by ManifestParser.

Shows every row as one line of an editable QTableWidget. Edits are persisted
immediately via CorrectionStore. Includes a filter to focus on rows with
empty key fields (the typical correction workflow).

All UI text is in French (per AGL mandate).
"""
from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Optional

from PyQt5.QtCore import Qt, QSignalBlocker, QThread, pyqtSignal
from PyQt5.QtGui import QColor, QBrush, QFont
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QCheckBox, QLineEdit, QFileDialog, QMessageBox,
    QHeaderView, QComboBox, QPlainTextEdit, QListWidget, QListWidgetItem,
    QGroupBox, QFormLayout, QFrame, QSplitter, QScrollArea,
)

from ..corrections import CorrectionStore
from ..exporter import ExcelExporter
from ..config import EXPORTS_DIR
from ..midas_mapper import MIDAS_COLUMNS, map_rows_to_midas
from ..ai import validate_rows, has_api_key
from .ai_dialogs import AIFixWorker, ensure_api_key


# Read-only metadata columns (never editable)
READONLY_FIELDS = {"source_file", "page", "_user_edited"}

# Key fields used by the "incomplete rows" filter
KEY_FIELDS_FOR_FILTER = (
    "bl_number", "container_number", "shipper", "consignee",
    "port_of_loading", "port_of_discharge", "weight",
)


class ManifestReviewDialog(QDialog):
    """Editable review dialog for ManifestParser output."""

    def __init__(
        self,
        rows: List[Dict[str, Any]],
        source_path: Path,
        parent=None,
        *,
        active_format: Optional[Dict] = None,
    ):
        super().__init__(parent)
        self.source_path = source_path
        # The learned format that produced these rows (if any). Used by the
        # feedback dialog to attribute feedback without re-detection.
        self.active_format = active_format
        self.store = CorrectionStore(source_path)
        # If we already have stored corrections for this doc, prefer them
        # over the freshly-parsed rows (user has edited them previously).
        if self.store.has_manifest_rows():
            stored = self.store.get_manifest_rows()
            # If the stored count matches the parsed count, use stored.
            # Otherwise rebuild from parsed (schema/version drift).
            if len(stored) == len(rows):
                self.rows = stored
            else:
                self.rows = list(rows)
                self.store.save_manifest_rows(self.rows)
        else:
            self.rows = list(rows)
            self.store.save_manifest_rows(self.rows)

        self.setWindowTitle(
            f"Revue du manifeste — {len(self.rows)} conteneurs ({source_path.name})"
        )
        self.resize(1400, 750)

        self._midas_mode = False  # False = vue brute éditable, True = aperçu MIDAS éditable
        self._midas_rows: Optional[List[Dict[str, Any]]] = None  # MIDAS rows with user edits
        self._group_by_bl = False  # True = collapse rows per bl_number (read-only)

        self._build_ui()
        self._populate_table()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------
    def _build_ui(self):
        lay = QVBoxLayout(self)

        # Header info
        info = QLabel(
            f"<b>{len(self.rows)}</b> conteneurs extraits de <b>{self.source_path.name}</b>."
            f"Toutes les modifications sont <b>enregistrées automatiquement</b>."
        )
        info.setStyleSheet("padding: 6px;")
        lay.addWidget(info)

        # Filter bar
        bar = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Rechercher (BL, conteneur, expéditeur…)")
        self.search.textChanged.connect(self._refilter)
        bar.addWidget(self.search, 3)

        self.chk_incomplete = QCheckBox("Afficher seulement les lignes incomplètes")
        self.chk_incomplete.toggled.connect(self._refilter)
        bar.addWidget(self.chk_incomplete, 1)

        self.chk_edited = QCheckBox("Lignes modifiées seulement")
        self.chk_edited.toggled.connect(self._refilter)
        bar.addWidget(self.chk_edited, 1)

        self.chk_group_bl = QCheckBox("Grouper par BL")
        self.chk_group_bl.setToolTip(
            "Affiche une ligne par connaissement (BL) au lieu d'une ligne par conteneur.\n"
            "Les conteneurs et poids sont agrégés. Vue en lecture seule —\n"
            "désactivez pour éditer les détails par conteneur."
        )
        self.chk_group_bl.toggled.connect(self._toggle_group_bl)
        bar.addWidget(self.chk_group_bl, 1)

        lay.addLayout(bar)

        # Status label
        self.lbl_status = QLabel()
        self.lbl_status.setStyleSheet("color: #555; padding: 2px 6px;")
        lay.addWidget(self.lbl_status)

        # Main table
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed)
        self.table.itemChanged.connect(self._on_item_changed)
        lay.addWidget(self.table, 1)

        # Bottom bar
        btns = QHBoxLayout()
        self.btn_reset_row = QPushButton("↩ Réinitialiser la ligne sélectionnée")
        self.btn_reset_row.clicked.connect(self._reset_selected_row)
        btns.addWidget(self.btn_reset_row)

        self.btn_toggle_midas = QPushButton("Aperçu MIDAS")
        self.btn_toggle_midas.setCheckable(True)
        self.btn_toggle_midas.setToolTip(
            "Bascule entre la vue brute (éditable) et l'aperçu MIDAS 43 colonnes (lecture seule).\n"
            "L'aperçu MIDAS montre exactement ce qui sera exporté."
        )
        self.btn_toggle_midas.toggled.connect(self._toggle_midas_view)
        btns.addWidget(self.btn_toggle_midas)

        # AI quality checks
        self.btn_validate_ai = QPushButton("Vérifier la qualité")
        self.btn_validate_ai.setProperty("agl", "primary")
        self.btn_validate_ai.setToolTip(
            "Détecte les lignes incomplètes ou suspectes (n° BL/conteneur invalide,\n"
            "poids non numérique, champs obligatoires manquants)."
        )
        self.btn_validate_ai.clicked.connect(self._run_quality_check)
        btns.addWidget(self.btn_validate_ai)

        self.btn_ai_fix = QPushButton("Corriger les lignes problématiques avec l'IA")
        self.btn_ai_fix.setProperty("agl", "accent")
        self.btn_ai_fix.setToolTip(
            "Envoie chaque ligne problématique à Gemini avec le contexte du document\n"
            "pour correction automatique. Nécessite une clé API."
        )
        self.btn_ai_fix.clicked.connect(self._ai_fix_problems)
        btns.addWidget(self.btn_ai_fix)

        btns.addStretch(1)

        self.btn_export = QPushButton("Export brut")
        self.btn_export.setToolTip("Exporte les colonnes brutes (debug / inspection)")
        self.btn_export.clicked.connect(self._export_excel)
        btns.addWidget(self.btn_export)

        self.btn_export_midas = QPushButton("Export MIDAS")
        self.btn_export_midas.setProperty("agl", "gold")
        self.btn_export_midas.setToolTip("Exporte au format MIDAS 43 colonnes pour saisie d'intégration")
        self.btn_export_midas.clicked.connect(self._export_midas)
        btns.addWidget(self.btn_export_midas)

        self.btn_feedback = QPushButton("Feedback IA…")
        self.btn_feedback.setProperty("agl", "ghost")
        self.btn_feedback.setToolTip(
            "Envoyer un commentaire libre sur la qualité de l'extraction.\n"
            "Le feedback sera injecté dans les prochains apprentissages et\n"
            "extractions IA pour ce format."
        )
        self.btn_feedback.clicked.connect(self._send_feedback)
        btns.addWidget(self.btn_feedback)

        self.btn_close = QPushButton("Fermer")
        self.btn_close.clicked.connect(self.accept)
        btns.addWidget(self.btn_close)
        lay.addLayout(btns)

    # ------------------------------------------------------------------
    # Population
    # ------------------------------------------------------------------
    def _columns(self) -> List[str]:
        if not self.rows:
            return []
        # Build column order from first row, but exclude "_user_edited"
        cols = [k for k in self.rows[0].keys() if k != "_user_edited"]
        return cols

    def _populate_table(self):
        # MIDAS view takes precedence — when MIDAS is on, group_by_bl just
        # changes how MIDAS aggregates source rows (per-container vs per-BL).
        if self._midas_mode:
            self._populate_midas()
            return
        if self._group_by_bl:
            self._populate_grouped()
            return
        cols = self._columns()
        self.table.clear()
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setRowCount(len(self.rows))

        with QSignalBlocker(self.table):
            for r, row in enumerate(self.rows):
                for c, key in enumerate(cols):
                    val = row.get(key, "")
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    if key in READONLY_FIELDS:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setForeground(QBrush(QColor("#888")))
                    if row.get("_user_edited"):
                        item.setBackground(QBrush(QColor("#FFF8C5")))
                    self.table.setItem(r, c, item)

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        for c in range(len(cols)):
            self.table.resizeColumnToContents(c)
            # cap very wide columns
            if self.table.columnWidth(c) > 280:
                self.table.setColumnWidth(c, 280)

        self._refresh_status()

    def _populate_midas(self):
        """Render rows in MIDAS 43-column layout (editable).

        Empty mandatory cells are highlighted in red so the user can
        instantly spot extraction failures. Edits update self._midas_rows.

        When ``self._group_by_bl`` is True, the source rows are first
        aggregated per BL (one MIDAS line per connaissement instead of
        per container).
        """
        # (Re)compute if first time or rows changed
        if self._midas_rows is None:
            if self._group_by_bl:
                source_rows = self._build_grouped_rows()
                # _build_grouped_rows leaves internal _pack_sum etc. — strip
                source_rows = [
                    {k: v for k, v in r.items() if not str(k).startswith("_")}
                    for r in source_rows
                ]
            else:
                source_rows = self.rows
            self._midas_rows = map_rows_to_midas(source_rows)
        midas_rows = self._midas_rows
        cols = MIDAS_COLUMNS
        self.table.clear()
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels(cols)
        self.table.setRowCount(len(midas_rows))

        # Colonnes laissées volontairement vides (saisie équipe d'intégration)
        intentionally_empty = {
            "Numéro escale", "Index", "Range",
            "Code transitaire", "chargeur.code", "Code marchandise",
            "Manutentionaire", "Port transbo1", "Port transbo2",
        }
        red_brush = QBrush(QColor("#FFD6D6"))
        gray_brush = QBrush(QColor("#EEEEEE"))

        with QSignalBlocker(self.table):
            for r, row in enumerate(midas_rows):
                for c, key in enumerate(cols):
                    val = row.get(key, "")
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    if not str(val).strip():
                        if key in intentionally_empty:
                            item.setBackground(gray_brush)
                            item.setToolTip("Saisie équipe d'intégration")
                        else:
                            item.setBackground(red_brush)
                            item.setToolTip("Champ vide — extraction à vérifier")
                    self.table.setItem(r, c, item)

        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        for c in range(len(cols)):
            self.table.resizeColumnToContents(c)
            if self.table.columnWidth(c) > 240:
                self.table.setColumnWidth(c, 240)

        self._refresh_status()

    def _toggle_midas_view(self, checked: bool):
        self._midas_mode = checked
        self.btn_toggle_midas.setText("Vue brute (édition)" if checked else "Aperçu MIDAS")
        # Force MIDAS recomputation (group state may have changed since last view).
        self._midas_rows = None
        self._populate_table()
        if checked:
            mode = "groupé par BL" if self._group_by_bl else "par conteneur"
            n_lines = len(self._midas_rows or [])
            self.lbl_status.setText(
                f"Aperçu MIDAS ({mode}) — {n_lines} lignes × {len(MIDAS_COLUMNS)} colonnes"
                f" · rouge = vide à corriger · gris = saisie équipe"
            )

    def _toggle_group_bl(self, checked: bool):
        self._group_by_bl = checked
        # If MIDAS is also on, recompute MIDAS from the new (grouped/un-grouped) rows.
        if self._midas_mode:
            self._midas_rows = None
        self._populate_table()
        if self._midas_mode:
            mode = "groupé par BL" if checked else "par conteneur"
            n_lines = len(self._midas_rows or [])
            self.lbl_status.setText(
                f"Aperçu MIDAS ({mode}) — {n_lines} lignes × {len(MIDAS_COLUMNS)} colonnes"
            )

    def _build_grouped_rows(self) -> List[Dict[str, Any]]:
        """Aggregate self.rows by bl_number — one row per BL.

        - container_number / seal1 → comma-joined uniques
        - weight / pack_qty / volume → numeric sum
        - text fields → first non-empty
        - container_count → injected
        """
        from collections import OrderedDict
        groups: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()

        def _to_float(v) -> float:
            try:
                s = str(v or "").replace(",", "").replace(" ", "")
                return float(s) if s else 0.0
            except Exception:
                return 0.0

        for row in self.rows:
            bl = str(row.get("bl_number") or "").strip() or "(BL inconnu)"
            g = groups.get(bl)
            if g is None:
                g = {
                    "bl_number": bl,
                    "container_count": 0,
                    "containers": [],
                    "seals": [],
                    "_weight_sum": 0.0,
                    "_pack_sum": 0.0,
                    "_volume_sum": 0.0,
                }
                groups[bl] = g
            g["container_count"] += 1
            cn = str(row.get("container_number") or "").strip()
            if cn and cn not in g["containers"]:
                g["containers"].append(cn)
            sl = str(row.get("seal1") or "").strip()
            if sl and sl not in g["seals"]:
                g["seals"].append(sl)
            g["_weight_sum"] += _to_float(row.get("weight"))
            g["_pack_sum"] += _to_float(row.get("pack_qty"))
            g["_volume_sum"] += _to_float(row.get("volume"))
            # Auto-pick first non-empty for descriptive fields
            for k in (
                "shipper", "consignee", "notify", "freight_forwarder",
                "port_of_loading", "port_of_discharge",
                "place_of_acceptance", "place_of_delivery",
                "weight_unit", "pack_unit", "volume_unit",
                "description", "bl_type",
            ):
                if not g.get(k):
                    v = str(row.get(k) or "").strip()
                    if v:
                        g[k] = v

        # Finalize: stringify aggregates
        out: List[Dict[str, Any]] = []
        for g in groups.values():
            g["container_number"] = ", ".join(g.pop("containers"))
            g["seal1"] = ", ".join(g.pop("seals"))
            ws = g.pop("_weight_sum")
            g["weight"] = f"{ws:.2f}".rstrip("0").rstrip(".") if ws else ""
            ps = g.pop("_pack_sum")
            g["pack_qty"] = str(int(ps)) if ps else ""
            vs = g.pop("_volume_sum")
            g["volume"] = f"{vs:.3f}".rstrip("0").rstrip(".") if vs else ""
            out.append(g)
        return out

    def _populate_grouped(self):
        rows = self._build_grouped_rows()
        cols = [
            "bl_number", "container_count", "container_number", "seal1",
            "shipper", "consignee", "notify",
            "port_of_loading", "port_of_discharge",
            "place_of_acceptance", "place_of_delivery",
            "weight", "weight_unit", "pack_qty", "pack_unit",
            "volume", "volume_unit", "description",
        ]
        labels = {
            "bl_number": "BL", "container_count": "Nb conteneurs",
            "container_number": "Conteneurs", "seal1": "Scellés",
            "shipper": "Expéditeur", "consignee": "Destinataire",
            "notify": "Notify",
            "port_of_loading": "Port chargement",
            "port_of_discharge": "Port déchargement",
            "place_of_acceptance": "Lieu prise en charge",
            "place_of_delivery": "Lieu livraison",
            "weight": "Poids total", "weight_unit": "Unité",
            "pack_qty": "Colis", "pack_unit": "Type colis",
            "volume": "Volume", "volume_unit": "Unité vol.",
            "description": "Description",
        }
        self.table.clear()
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels([labels[c] for c in cols])
        self.table.setRowCount(len(rows))
        gold = QBrush(QColor("#FFF7E0"))
        with QSignalBlocker(self.table):
            for r, row in enumerate(rows):
                for c, key in enumerate(cols):
                    val = row.get(key, "")
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    if key == "bl_number":
                        item.setBackground(gold)
                        item.setForeground(QBrush(QColor("#1A4076")))
                        f = item.font(); f.setBold(True); item.setFont(f)
                    self.table.setItem(r, c, item)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Interactive)
        for c in range(len(cols)):
            self.table.resizeColumnToContents(c)
            if self.table.columnWidth(c) > 280:
                self.table.setColumnWidth(c, 280)
        self.lbl_status.setText(
            f"Vue groupée par BL — {len(rows)} BL · {len(self.rows)} conteneurs · lecture seule"
        )

    def _refresh_status(self):
        edited = sum(1 for r in self.rows if r.get("_user_edited"))
        incomplete = sum(1 for r in self.rows if self._is_incomplete(r))
        self.lbl_status.setText(
            f"{len(self.rows)} lignes ·"
            f"{edited} modifiées ·"
            f"{incomplete} incomplètes"
        )

    @staticmethod
    def _is_incomplete(row: Dict[str, Any]) -> bool:
        return any(not str(row.get(k) or "").strip() for k in KEY_FIELDS_FOR_FILTER if k in row)

    # ------------------------------------------------------------------
    # Edit handling
    # ------------------------------------------------------------------
    def _on_item_changed(self, item: QTableWidgetItem):
        r = item.row()
        c = item.column()
        new_val = item.text()

        # Grouped view is read-only — ignore any spurious signals
        if self._group_by_bl:
            return

        if self._midas_mode:
            # Edit in MIDAS view — update _midas_rows only (no CorrectionStore)
            if self._midas_rows is None or r >= len(self._midas_rows):
                return
            key = MIDAS_COLUMNS[c] if c < len(MIDAS_COLUMNS) else None
            if not key:
                return
            if str(self._midas_rows[r].get(key, "")) == new_val:
                return
            self._midas_rows[r][key] = new_val
            # Remove red highlight once user fills a cell
            with QSignalBlocker(self.table):
                item.setBackground(QBrush(QColor("#FFF8C5")))
                item.setToolTip("Modifié manuellement")
            self._refresh_status()
            return

        cols = self._columns()
        if c >= len(cols):
            return
        key = cols[c]
        if key in READONLY_FIELDS:
            return
        new_val = item.text()
        old_val = self.rows[r].get(key, "")
        if str(old_val) == new_val:
            return
        # Persist
        self.rows[r][key] = new_val
        self.store.update_manifest_row(r, {key: new_val})
        # Visual flag
        with QSignalBlocker(self.table):
            for cc in range(self.table.columnCount()):
                it = self.table.item(r, cc)
                if it:
                    it.setBackground(QBrush(QColor("#FFF8C5")))
        self._refresh_status()

    # ------------------------------------------------------------------
    # Quality validation + AI auto-fix
    # ------------------------------------------------------------------
    def _run_quality_check(self):
        """Highlight rows with detected issues. Updates row tooltip + status bar."""
        if self._midas_mode:
            self.btn_toggle_midas.setChecked(False)  # switch back to raw view
        issues = validate_rows(self.rows)
        # Reset all backgrounds in raw view
        if not self._midas_mode:
            self._populate_table()
        # Apply red tint + tooltip on problematic rows
        red = QBrush(QColor("#FFD6D6"))
        for ridx, iss in issues.items():
            if ridx >= self.table.rowCount():
                continue
            tip = "·".join(iss)
            for c in range(self.table.columnCount()):
                it = self.table.item(ridx, c)
                if it:
                    if not self.rows[ridx].get("_user_edited"):
                        it.setBackground(red)
                    it.setToolTip(tip)
        self.lbl_status.setText(
            f"{len(self.rows)} lignes · {len(issues)} avec problèmes détectés"
            + ("· cliquez « Corriger avec IA »" if issues else "")
        )
        if not issues:
            QMessageBox.information(self, "Qualité OK", "Aucun problème détecté.")

    def _ai_fix_problems(self):
        """Send each problematic row to Gemini for correction."""
        if not ensure_api_key(self):
            return
        issues_map = validate_rows(self.rows)
        if not issues_map:
            QMessageBox.information(self, "Rien à corriger",
                                    "Aucun problème détecté. Lancez d'abord « Vérifier la qualité ».")
            return
        if QMessageBox.question(
            self, "Correction IA",
            f"<b>{len(issues_map)} ligne(s)</b> seront envoyées à Gemini pour correction.<br><br>"
            "Cela peut prendre quelques secondes par ligne. Continuer ?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes,
        ) != QMessageBox.Yes:
            return

        # Build a context string from the source PDF (text-only, best effort)
        context = self._build_doc_context()

        # Process sequentially via a small queue of workers
        self._fix_queue = list(issues_map.items())
        self._fix_total = len(self._fix_queue)
        self._fix_done = 0
        self._fix_context = context
        self.btn_ai_fix.setEnabled(False)
        self.lbl_status.setText(f"Correction IA en cours… 0 / {self._fix_total}")
        self._launch_next_fix()

    def _launch_next_fix(self):
        if not self._fix_queue:
            self.btn_ai_fix.setEnabled(True)
            self.lbl_status.setText(
                f"Correction IA terminée — {self._fix_done} / {self._fix_total} ligne(s) traitée(s)."
            )
            self._populate_table()
            return
        ridx, iss = self._fix_queue.pop(0)
        self._cur_fix_worker = AIFixWorker(ridx, dict(self.rows[ridx]), iss, self._fix_context)
        self._cur_fix_worker.finished_ok.connect(self._on_fix_done)
        self._cur_fix_worker.failed.connect(self._on_fix_failed)
        self._cur_fix_worker.start()

    def _on_fix_done(self, ridx: int, fixed: dict):
        # Apply only changed values, mark row as edited, persist
        original = self.rows[ridx]
        changed = False
        for k, v in fixed.items():
            if k in ("source_file", "page", "_user_edited"):
                continue
            if str(v).strip() and str(original.get(k, "")).strip() != str(v).strip():
                original[k] = str(v).strip()
                changed = True
        if changed:
            original["_user_edited"] = True
            self.store.save_manifest_rows(self.rows)
        self._fix_done += 1
        self.lbl_status.setText(
            f"Correction IA en cours… {self._fix_done} / {self._fix_total}"
        )
        self._launch_next_fix()

    def _on_fix_failed(self, ridx: int, err: str):
        self._fix_done += 1
        # Keep going — don't abort the whole batch on a single failure
        self._launch_next_fix()

    def _build_doc_context(self) -> str:
        """Best-effort: extract embedded text from the source PDF (max ~16k chars)."""
        try:
            import pdfplumber
            chunks = []
            total = 0
            with pdfplumber.open(str(self.source_path)) as pdf:
                for p in pdf.pages:
                    t = p.extract_text() or ""
                    if t:
                        chunks.append(t)
                        total += len(t)
                        if total > 16000:
                            break
            return "\n\n".join(chunks)[:16000]
        except Exception:
            return ""

    def _reset_selected_row(self):
        rows = sorted({i.row() for i in self.table.selectedIndexes()})
        if not rows:
            return
        if QMessageBox.question(
            self, "Réinitialiser",
            f"Réinitialiser {len(rows)} ligne(s) sélectionnée(s) ?\n"
            f"(Le drapeau « modifiée » sera retiré, les valeurs gardées en l'état)",
            QMessageBox.Yes | QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        for r in rows:
            self.rows[r].pop("_user_edited", None)
            self.store.data.manifest_rows[r].pop("_user_edited", None)
        self.store._flush()
        self._populate_table()

    # ------------------------------------------------------------------
    # Filter
    # ------------------------------------------------------------------
    def _refilter(self):
        # In grouped/MIDAS views the table row count doesn't match self.rows
        if self._group_by_bl or self._midas_mode:
            return
        text = self.search.text().lower().strip()
        only_incomplete = self.chk_incomplete.isChecked()
        only_edited = self.chk_edited.isChecked()
        for r, row in enumerate(self.rows):
            visible = True
            if text:
                blob = "".join(str(v) for v in row.values()).lower()
                if text not in blob:
                    visible = False
            if visible and only_incomplete and not self._is_incomplete(row):
                visible = False
            if visible and only_edited and not row.get("_user_edited"):
                visible = False
            self.table.setRowHidden(r, not visible)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def _export_excel(self):
        if not self.rows:
            QMessageBox.information(self, "Rien à exporter", "Aucune ligne disponible.")
            return
        default = str(EXPORTS_DIR / f"{self.source_path.stem}_export.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter vers Excel", default, "Fichiers Excel (*.xlsx)"
        )
        if not path:
            return
        # Strip internal flags before export
        clean = [{k: v for k, v in r.items() if not k.startswith("_")} for r in self.rows]
        out = ExcelExporter(output_path=Path(path)).export(clean)
        QMessageBox.information(
            self, "Export terminé",
            f"{len(clean)} lignes exportées vers :\n{out}"
        )

    def _send_feedback(self):
        """Open the rich feedback dialog (auto-detects format, attaches
        rows + page snapshots, pre-flags rows the user has selected).
        """
        try:
            from ..ai.format_registry import detect_learned, list_learned
        except Exception:
            QMessageBox.warning(self, "Indisponible",
                                "Le module IA n'est pas disponible.")
            return

        # Resolve the target format with this priority:
        #   1. The format actually used to produce these rows (passed in
        #      via active_format) — most accurate.
        #   2. Auto-detection from the source PDF text.
        # Falls back to the user picking one in the dialog if both fail.
        learned = self.active_format
        if not learned:
            sample = self._build_doc_context()
            learned = detect_learned(sample) if sample else None

        # Pre-flag any rows the user has explicitly selected in the table
        # — those become the "problem rows" of the feedback by default.
        try:
            sel_rows = sorted({i.row() for i in self.table.selectedIndexes()})
        except Exception:
            sel_rows = []
        # Filter to actual data rows (not grouped / midas views)
        if self._group_by_bl or self._midas_mode:
            sel_rows = []  # row indexes don't map to self.rows in those views

        all_formats = []
        try:
            all_formats = list_learned()
        except Exception:
            pass
        if not all_formats:
            QMessageBox.information(
                self, "Aucun format appris",
                "Aucun format n'a encore été appris. Apprenez d'abord ce "
                "format à l'IA puis donnez du feedback ici."
            )
            return

        dlg = FeedbackDialog(
            parent=self,
            source_path=self.source_path,
            rows=self.rows,
            problem_indexes=sel_rows,
            detected_format=learned,
            all_formats=all_formats,
        )
        if dlg.exec_() == QDialog.Accepted:
            self.lbl_status.setText(
                f"Feedback envoyé au format « {dlg.target_name} »."
            )
            if getattr(dlg, "_relaunch_learn", False):
                # Bubble the request up to MainWindow which owns the
                # learn worker. Close the review dialog so the user sees
                # the new pipeline run.
                parent = self.parent()
                if parent is not None and hasattr(parent, "_ai_learn_format"):
                    self.accept()
                    parent._ai_learn_format()

    def _export_midas(self):
        """Export at MIDAS 43-column format, using edited MIDAS rows if available."""
        if not self.rows:
            QMessageBox.information(self, "Rien à exporter", "Aucune ligne disponible.")
            return
        default = str(EXPORTS_DIR / f"MIDAS_{self.source_path.stem}.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Exporter au format MIDAS", default, "Fichiers Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            exp = ExcelExporter(output_path=Path(path))
            if self._midas_rows is not None:
                # User has edited the MIDAS view — export directly from the
                # edited MIDAS rows (already in 43-column format).
                from ..midas_mapper import MIDAS_COLUMNS as _COLS
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = Workbook()
                ws = wb.active
                ws.title = "MIDAS"
                ws.append(_COLS)
                header_font = Font(bold=True, color="FFFFFF", size=10)
                header_fill = PatternFill("solid", fgColor="1A4076")
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                ws.row_dimensions[1].height = 32
                for row in self._midas_rows:
                    ws.append([row.get(h, "") for h in _COLS])
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                for col_cells in ws.columns:
                    vals = [c.value for c in col_cells if c.value is not None]
                    length = max((len(str(v)) for v in vals), default=10)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 32)
                wb.save(path)
                out = Path(path)
            else:
                out = exp.export_midas(self.rows)
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'export MIDAS", str(e))
            return
        QMessageBox.information(
            self, "Export MIDAS terminé",
            f"{len(self._midas_rows or self.rows)} ligne(s) exportée(s) au format MIDAS vers :\n{out}\n\n"
            f"Colonnes laissées vides (saisie équipe d'intégration) :\n"
            f"• Numéro escale, Index, Range\n"
            f"• Code transitaire / chargeur / marchandise\n"
            f"• Manutentionaire",
        )


# ════════════════════════════════════════════════════════════════════════
# Background page-render worker for the feedback dialog
# ════════════════════════════════════════════════════════════════════════
class _PageRenderWorker(QThread):
    """Renders PDF pages to PNG one by one in the background.

    Emits ``page_ready(pg_num, png_bytes)`` for each successfully rendered
    page. Allows the FeedbackDialog to open instantly while pages stream
    in — no more "joindre la page 1 only" because we skipped the wait.
    """

    page_ready = pyqtSignal(int, bytes)
    page_failed = pyqtSignal(int, str)
    all_done = pyqtSignal()

    def __init__(self, pdf_path: str, pages: List[int], dpi: int = 110):
        super().__init__()
        self.pdf_path = pdf_path
        self.pages = list(pages)
        self.dpi = dpi
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        try:
            from pdf2image import convert_from_path
            from ..paths import poppler_bin
            import io as _io
            poppler = poppler_bin()
        except Exception as e:
            for p in self.pages:
                self.page_failed.emit(p, str(e))
            self.all_done.emit()
            return

        for pg_num in self.pages:
            if self._cancel:
                break
            try:
                images = convert_from_path(
                    self.pdf_path,
                    dpi=self.dpi,
                    poppler_path=poppler,
                    fmt="png",
                    first_page=pg_num,
                    last_page=pg_num,
                )
                if not images:
                    self.page_failed.emit(pg_num, "page introuvable")
                    continue
                buf = _io.BytesIO()
                images[0].save(buf, format="PNG", optimize=True)
                self.page_ready.emit(pg_num, buf.getvalue())
            except Exception as e:
                self.page_failed.emit(pg_num, str(e))
        self.all_done.emit()


# ════════════════════════════════════════════════════════════════════════
# Feedback dialog — auto-detect format, attach rows + page snapshots
# ════════════════════════════════════════════════════════════════════════
class FeedbackDialog(QDialog):
    """Rich feedback dialog for the AI learning loop.

    Features:
      * Auto-detects the active learned format (no manual picker unless
        detection fails AND multiple formats exist).
      * Pre-flags rows the user has selected in the review table as the
        "problem rows" — those are highlighted in a compact preview.
      * Auto-attaches page snapshots (PNG) of the source PDF — the user
        can pick which pages to include and (in a future iteration)
        annotate problem zones.
      * Persists everything via :func:`format_registry.add_feedback`.
    """

    def __init__(
        self,
        *,
        parent,
        source_path: Path,
        rows: List[Dict[str, Any]],
        problem_indexes: List[int],
        detected_format: Optional[Dict],
        all_formats: List[Dict],
    ):
        super().__init__(parent)
        from PyQt5.QtGui import QPixmap
        self._QPixmap = QPixmap

        self.source_path = Path(source_path)
        self.rows = rows or []
        self.problem_indexes = list(problem_indexes or [])
        self.all_formats = all_formats or []
        # Resolve target format: detected → first one. Always changeable.
        if detected_format and detected_format.get("name"):
            self.target_name = detected_format["name"]
            self._auto_detected = True
        else:
            self.target_name = (self.all_formats[0].get("name") or "")
            self._auto_detected = False

        self._page_snapshots: List[Dict] = []  # [{path, pixmap, label, include}]
        self._image_paths_to_save: List[bytes] = []  # raw png bytes saved on accept
        self._relaunch_learn: bool = False  # set by _on_send if user wants re-learn

        self.setWindowTitle("Feedback IA — apprentissage continu")
        self.resize(1100, 720)

        self._build_ui()
        self._render_page_snapshots()

    # ------------------------------------------------------------------
    def _build_ui(self):
        root = QVBoxLayout(self)

        # ── Header: format target (auto-detected, overridable) ───────
        header = QGroupBox("Format ciblé")
        hf = QFormLayout(header)
        self.combo_format = QComboBox()
        names = [f.get("name", "") for f in self.all_formats if f.get("name")]
        for n in names:
            self.combo_format.addItem(n)
        if self.target_name in names:
            self.combo_format.setCurrentIndex(names.index(self.target_name))
        self.combo_format.currentTextChanged.connect(self._on_format_changed)
        hf.addRow("Format :", self.combo_format)
        self.lbl_detect = QLabel()
        self.lbl_detect.setStyleSheet("color: #1A4076;" if self._auto_detected else "color: #A87E1F;")
        self.lbl_detect.setText(
            "✓ Format actif (utilisé pour produire l'extraction courante)."
            if self._auto_detected else
            "⚠ Aucun format actif identifié — choix par défaut. Ajustez si nécessaire."
        )
        hf.addRow("", self.lbl_detect)
        root.addWidget(header)

        # ── Splitter: left = comment + problem rows, right = page snapshots ──
        split = QSplitter(Qt.Horizontal)
        root.addWidget(split, 1)

        # ── LEFT panel ────────────────────────────────────────────────
        left = QFrame()
        ll = QVBoxLayout(left)

        # Comment box
        ll.addWidget(QLabel("<b>Décrivez les erreurs ou améliorations souhaitées :</b>"))
        self.txt_comment = QPlainTextEdit()
        self.txt_comment.setPlaceholderText(
            "Exemples :\n"
            "• Les colonnes Shipper et Consignee sont systématiquement inversées.\n"
            "• Le poids brut est en livres, il faut le convertir en kg.\n"
            "• Les lignes vides au milieu du tableau sont à ignorer."
        )
        self.txt_comment.setMinimumHeight(140)
        ll.addWidget(self.txt_comment, 1)

        # Problem rows list (pre-flagged from table selection)
        gb_rows = QGroupBox(
            f"Lignes problématiques ({len(self.problem_indexes)} sélectionnée(s))"
        )
        gv = QVBoxLayout(gb_rows)
        self.lst_problems = QListWidget()
        self.lst_problems.setSelectionMode(QListWidget.MultiSelection)
        self._populate_problem_list()
        gv.addWidget(self.lst_problems)
        # "Include all rows" toggle
        self.chk_include_all = QCheckBox(
            f"Joindre l'extraction complète ({len(self.rows)} ligne(s)) en plus des lignes flaguées"
        )
        gv.addWidget(self.chk_include_all)
        ll.addWidget(gb_rows, 1)

        split.addWidget(left)

        # ── RIGHT panel: page snapshots ──────────────────────────────
        right = QFrame()
        rl = QVBoxLayout(right)
        rl.addWidget(QLabel("<b>Captures de pages à joindre :</b>"))
        rl.addWidget(QLabel(
            "<i>Sélectionnez les pages qui montrent les erreurs. Les images "
            "sont jointes au feedback et envoyées à l'IA lors de la prochaine "
            "extraction.</i>"
        ))
        self.scroll_pages = QScrollArea()
        self.scroll_pages.setWidgetResizable(True)
        self._pages_container = QFrame()
        self._pages_layout = QVBoxLayout(self._pages_container)
        self.scroll_pages.setWidget(self._pages_container)
        rl.addWidget(self.scroll_pages, 1)

        split.addWidget(right)
        split.setStretchFactor(0, 3)
        split.setStretchFactor(1, 2)

        # ── Bottom buttons ────────────────────────────────────────────
        bb = QHBoxLayout()
        bb.addStretch(1)
        self.btn_cancel = QPushButton("Annuler")
        self.btn_cancel.clicked.connect(self.reject)
        bb.addWidget(self.btn_cancel)
        self.btn_send = QPushButton("Envoyer le feedback")
        self.btn_send.setProperty("agl", "gold")
        self.btn_send.setDefault(True)
        self.btn_send.clicked.connect(self._on_send)
        bb.addWidget(self.btn_send)
        root.addLayout(bb)

    # ------------------------------------------------------------------
    def _populate_problem_list(self):
        self.lst_problems.clear()
        if not self.problem_indexes:
            it = QListWidgetItem(
                "Aucune ligne sélectionnée — sélectionnez des lignes dans la "
                "vue principale avant d'ouvrir le feedback pour les flaguer "
                "automatiquement ici."
            )
            it.setFlags(Qt.NoItemFlags)
            self.lst_problems.addItem(it)
            return
        for idx in self.problem_indexes:
            if not (0 <= idx < len(self.rows)):
                continue
            r = self.rows[idx]
            label = (
                f"Ligne {idx + 1}  ·  "
                f"BL={r.get('bl_number', '?')}  ·  "
                f"Conteneur={r.get('container_number', '?')}  ·  "
                f"{(r.get('shipper') or '')[:30]} → {(r.get('consignee') or '')[:30]}"
            )
            it = QListWidgetItem(label)
            it.setData(Qt.UserRole, idx)
            it.setSelected(True)
            self.lst_problems.addItem(it)

    # ------------------------------------------------------------------
    def _on_format_changed(self, name: str):
        self.target_name = name
        self.lbl_detect.setText(
            "✓ Détecté automatiquement." if (self._auto_detected and name == (self.combo_format.itemText(0)))
            else "Choix manuel."
        )
        self.lbl_detect.setStyleSheet("color: #1A4076;")

    # ------------------------------------------------------------------
    def _render_page_snapshots(self):
        """Render PDF pages as PNG thumbnails in a background thread.

        Builds a placeholder card per page IMMEDIATELY (so the user can
        already check/uncheck them), then streams the actual pixmaps as
        the worker delivers them. This avoids the "I only joined page 1
        because I clicked Send before the renders finished" problem.
        """
        # 1) Determine total page count and which pages to expose.
        total_pages = self._probe_page_count()
        if total_pages <= 0:
            self._pages_layout.addWidget(QLabel(
                "<i>Impossible de lire le nombre de pages du PDF.</i>"
            ))
            self._pages_layout.addStretch(1)
            return

        # Pre-flag pages referenced by problem rows. Robust to rows
        # whose ``page`` field is missing / 0 / non-numeric.
        self._total_pages = total_pages
        prio_pages = self._resolve_page_numbers(self.problem_indexes)

        # Cap to a reasonable upper bound (the dialog stays usable).
        MAX_PAGES = 30
        all_pages = list(range(1, min(total_pages, MAX_PAGES) + 1))

        # 2) Build placeholder cards (instant UI).
        for pg_num in all_pages:
            default_checked = pg_num in prio_pages
            self._add_page_placeholder(pg_num, default_checked=default_checked)

        # Status banner shown while rendering.
        self._render_status = QLabel(
            f"<i>Chargement des aperçus en arrière-plan… (0/{len(all_pages)})</i>"
        )
        self._render_status.setStyleSheet("color: #1A4076; padding: 4px;")
        self._pages_layout.addWidget(self._render_status)
        self._pages_layout.addStretch(1)
        self._render_total = len(all_pages)
        self._render_done = 0

        # 3) Order: priority pages first, then the rest.
        ordered = sorted(all_pages, key=lambda p: (p not in prio_pages, p))

        # 4) Kick off the worker.
        self._page_worker = _PageRenderWorker(str(self.source_path), ordered)
        self._page_worker.page_ready.connect(self._on_page_rendered)
        self._page_worker.page_failed.connect(self._on_page_render_failed)
        self._page_worker.all_done.connect(self._on_all_pages_rendered)
        self._page_worker.start()

    # ------------------------------------------------------------------
    def _probe_page_count(self) -> int:
        """Return the total number of pages in the source PDF (or 0)."""
        try:
            from pdf2image.pdf2image import pdfinfo_from_path
            from ..paths import poppler_bin
            info = pdfinfo_from_path(str(self.source_path), poppler_path=poppler_bin())
            return int(info.get("Pages", 0) or 0)
        except Exception:
            pass
        # Fallback: try pdfplumber
        try:
            import pdfplumber
            with pdfplumber.open(str(self.source_path)) as pdf:
                return len(pdf.pages)
        except Exception:
            return 0

    # ------------------------------------------------------------------
    def _resolve_page_numbers(self, flagged_idxs):
        """Return a set of 1-based page numbers covering the flagged rows.

        Robust to rows whose ``page`` field is missing, ``0``, or non
        numeric. When unknown, estimate by interpolating the row's
        position in ``self.rows`` over the document's total page count.
        Always returns a non-empty set (defaults to ``{1}``).
        """
        pages = set()
        total_rows = max(1, len(self.rows))
        total_pages = max(1, int(getattr(self, "_total_pages", 0) or 0))
        for i in flagged_idxs or []:
            if not (0 <= i < len(self.rows)):
                continue
            raw = self.rows[i].get("page")
            try:
                pg = int(raw) if raw is not None else 0
            except (TypeError, ValueError):
                pg = 0
            if pg <= 0:
                # Estimate by row position.
                pg = max(1, round((i + 1) / total_rows * total_pages))
            pg = min(max(1, pg), total_pages)
            pages.add(pg)
        return pages or {1}

    # ------------------------------------------------------------------
    def _add_page_placeholder(self, pg_num: int, *, default_checked: bool):
        """Add an empty card with checkbox + "Chargement…" placeholder."""
        box = QFrame()
        box.setFrameShape(QFrame.StyledPanel)
        bv = QVBoxLayout(box)
        chk = QCheckBox(f"Joindre la page {pg_num}")
        chk.setChecked(default_checked)
        chk.setEnabled(False)  # re-enabled when the page is ready
        bv.addWidget(chk)
        thumb = QLabel("⏳ Chargement de l'aperçu…")
        thumb.setAlignment(Qt.AlignCenter)
        thumb.setMinimumHeight(120)
        thumb.setStyleSheet(
            "border: 1px dashed #CCC; color: #888; padding: 24px; "
            "background: #FAFAFA;"
        )
        bv.addWidget(thumb)
        # Insert BEFORE the trailing stretch / status banner.
        self._pages_layout.insertWidget(self._pages_layout.count(), box)
        self._page_snapshots.append({
            "page": pg_num,
            "checkbox": chk,
            "thumb": thumb,
            "png_bytes": None,  # filled in by _on_page_rendered
            "box": box,
        })

    # ------------------------------------------------------------------
    def _on_page_rendered(self, pg_num: int, png_bytes: bytes):
        from PyQt5.QtGui import QPixmap
        snap = next((s for s in self._page_snapshots if s["page"] == pg_num), None)
        if not snap:
            return
        pix = QPixmap()
        pix.loadFromData(png_bytes)
        thumb: QLabel = snap["thumb"]
        max_w = 480
        if pix.width() > max_w:
            thumb.setPixmap(pix.scaledToWidth(max_w, Qt.SmoothTransformation))
        else:
            thumb.setPixmap(pix)
        thumb.setStyleSheet("border: 1px solid #CCC;")
        thumb.setText("")
        snap["png_bytes"] = png_bytes
        snap["checkbox"].setEnabled(True)
        self._render_done += 1
        if hasattr(self, "_render_status") and self._render_status:
            self._render_status.setText(
                f"<i>Chargement des aperçus en arrière-plan… "
                f"({self._render_done}/{self._render_total})</i>"
            )

    # ------------------------------------------------------------------
    def _on_page_render_failed(self, pg_num: int, err: str):
        snap = next((s for s in self._page_snapshots if s["page"] == pg_num), None)
        if snap:
            snap["thumb"].setText(f"❌ Échec : {err[:80]}")
            snap["thumb"].setStyleSheet(
                "border: 1px solid #C44; color: #C44; padding: 24px;"
            )
            snap["checkbox"].setEnabled(False)
            snap["checkbox"].setChecked(False)
        self._render_done += 1
        if hasattr(self, "_render_status") and self._render_status:
            self._render_status.setText(
                f"<i>Chargement des aperçus… ({self._render_done}/{self._render_total})</i>"
            )

    # ------------------------------------------------------------------
    def _on_all_pages_rendered(self):
        if hasattr(self, "_render_status") and self._render_status:
            self._render_status.setText(
                f"<i>✓ Aperçus chargés ({self._render_done}/{self._render_total} pages).</i>"
            )
            self._render_status.setStyleSheet("color: #2E7D32; padding: 4px;")

    # ------------------------------------------------------------------
    def closeEvent(self, event):
        # Stop the background renderer if the user closes the dialog early.
        worker = getattr(self, "_page_worker", None)
        if worker and worker.isRunning():
            worker.cancel()
            worker.wait(2000)
        super().closeEvent(event)

    # ------------------------------------------------------------------
    def _add_page_thumb(self, pg_num: int, pix, png_bytes: bytes, *, default_checked: bool):
        from PyQt5.QtGui import QPixmap
        box = QFrame()
        box.setFrameShape(QFrame.StyledPanel)
        bv = QVBoxLayout(box)
        chk = QCheckBox(f"Joindre la page {pg_num}")
        chk.setChecked(default_checked)
        bv.addWidget(chk)
        thumb = QLabel()
        # Cap the displayed thumbnail size so the dialog doesn't explode
        # on full-resolution renders.
        max_w = 480
        if pix.width() > max_w:
            thumb.setPixmap(pix.scaledToWidth(max_w, Qt.SmoothTransformation))
        else:
            thumb.setPixmap(pix)
        thumb.setStyleSheet("border: 1px solid #CCC;")
        bv.addWidget(thumb)
        self._pages_layout.addWidget(box)
        self._page_snapshots.append({
            "page": pg_num,
            "checkbox": chk,
            "png_bytes": png_bytes,
        })

    # ------------------------------------------------------------------
    def _on_send(self):
        text = self.txt_comment.toPlainText().strip()
        if not text:
            QMessageBox.warning(
                self, "Commentaire requis",
                "Décrivez l'erreur ou l'amélioration souhaitée avant d'envoyer."
            )
            return
        self.target_name = (self.combo_format.currentText() or "").strip()
        if not self.target_name:
            QMessageBox.warning(self, "Format requis", "Sélectionnez un format cible.")
            return

        # Resolve flagged indexes from list selection
        flagged: List[int] = []
        for i in range(self.lst_problems.count()):
            it = self.lst_problems.item(i)
            if it is None or not (it.flags() & Qt.ItemIsSelectable):
                continue
            if it.isSelected():
                idx = it.data(Qt.UserRole)
                if isinstance(idx, int):
                    flagged.append(idx)

        # Determine the rows snapshot to attach
        if self.chk_include_all.isChecked():
            rows_snapshot = list(self.rows)
            problem_idx_in_snap = flagged
        elif flagged:
            rows_snapshot = [self.rows[i] for i in flagged if 0 <= i < len(self.rows)]
            problem_idx_in_snap = list(range(len(rows_snapshot)))
        else:
            rows_snapshot = []
            problem_idx_in_snap = []

        # Save selected page snapshots
        from ..ai.format_registry import add_feedback, save_feedback_image
        # NEW: structured spatial diffs replace screenshots as the
        # primary feedback signal. Screenshots are kept as a legacy
        # fallback (computed only if the user explicitly checked pages).
        diffs: List[dict] = []
        try:
            from ..ai.spatial_diff import compute_diffs
            sd = compute_diffs(self.rows, flagged, self.source_path)
            diffs = [d.to_dict() for d in sd]
        except Exception:
            diffs = []

        image_paths: List[str] = []
        pending_pages: List[int] = []
        for snap in self._page_snapshots:
            if not snap["checkbox"].isChecked():
                continue
            if not snap.get("png_bytes"):
                # Checked but the background renderer hasn't delivered this
                # page yet — record so we can warn the user.
                pending_pages.append(snap["page"])
                continue
            try:
                p = save_feedback_image(
                    self.target_name,
                    snap["png_bytes"],
                    label=f"page{snap['page']}",
                )
                image_paths.append(p)
            except Exception:
                continue

        # If the user clicked Send while some checked pages were still
        # rendering, give them a chance to wait instead of losing them.
        if pending_pages:
            pages_str = ", ".join(str(p) for p in pending_pages[:8])
            if len(pending_pages) > 8:
                pages_str += f" (+{len(pending_pages) - 8})"
            choice = QMessageBox.question(
                self, "Aperçus en cours de chargement",
                f"Les pages {pages_str} sont encore en cours de rendu.\n\n"
                "• Oui : attendre que ces aperçus soient prêts puis envoyer.\n"
                "• Non : envoyer maintenant sans ces images.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if choice == QMessageBox.Yes:
                # Wait for the worker, then re-collect the now-ready images.
                worker = getattr(self, "_page_worker", None)
                if worker and worker.isRunning():
                    worker.wait(60_000)
                for snap in self._page_snapshots:
                    if (snap["checkbox"].isChecked()
                            and snap.get("png_bytes")
                            and snap["page"] in pending_pages):
                        try:
                            p = save_feedback_image(
                                self.target_name,
                                snap["png_bytes"],
                                label=f"page{snap['page']}",
                            )
                            image_paths.append(p)
                        except Exception:
                            continue

        ok = add_feedback(
            self.target_name,
            text,
            doc_name=self.source_path.name,
            rows_snapshot=rows_snapshot or None,
            problem_indexes=problem_idx_in_snap or None,
            image_paths=image_paths or None,
            diffs=diffs or None,
        )
        if not ok:
            QMessageBox.warning(
                self, "Échec",
                f"Impossible d'enregistrer le feedback pour « {self.target_name} »."
            )
            return
        # Offer immediate re-learning so the user sees the effect right away.
        msg = QMessageBox(self)
        msg.setWindowTitle("Feedback enregistré")
        msg.setIcon(QMessageBox.Information)
        msg.setText(
            f"Feedback ajouté au format « {self.target_name} ».<br>"
            f"  • Lignes flaguées : <b>{len(problem_idx_in_snap)}</b><br>"
            f"  • Images jointes : <b>{len(image_paths)}</b>"
        )
        msg.setInformativeText(
            "Voulez-vous <b>relancer l'apprentissage maintenant</b> ? "
            "Le feedback sera injecté dans le prompt et un nouveau parser sera généré."
        )
        btn_relearn = msg.addButton("Relancer l'apprentissage", QMessageBox.AcceptRole)
        btn_close = msg.addButton("Plus tard", QMessageBox.RejectRole)
        msg.exec_()
        self._relaunch_learn = (msg.clickedButton() is btn_relearn)
        self.accept()
